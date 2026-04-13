"""
Extract all cabinet bid and job cost data from MAC's bid sheet + P:/ drive bids.
Outputs structured JSON to G:\My Drive\MAC\Estimator\historical_data.json
"""
import openpyxl
import json
import os
import re
from datetime import datetime, date
from collections import defaultdict

BID_SHEET = r"C:\Users\Dylan\Downloads\Cabinet Shop Bid Sheet.xlsx"
BIDS_DIR = r"P:\Mac2026\Cabinets\Bids"
OUTPUT = r"G:\My Drive\MAC\Estimator\historical_data.json"


def parse_date(v):
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    return None


def safe_float(v):
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


# ── 1. Parse bid tracking sheets (2019-2026) ──

def parse_2026(ws):
    """4-column layout: Rusty Internal, Rusty Third Party, Josh Internal, Josh Third Party"""
    bids = []
    sections = [
        {"estimator": "Rusty", "type": "Internal", "name_col": "B", "date_col": "C", "bid_col": "D", "contract_col": "E"},
        {"estimator": "Rusty", "type": "Third Party", "name_col": "H", "date_col": "I", "bid_col": "J", "contract_col": "K"},
        {"estimator": "Josh", "type": "Internal", "name_col": "N", "date_col": "O", "bid_col": "P", "contract_col": "Q"},
        {"estimator": "Josh", "type": "Third Party", "name_col": "T", "date_col": "U", "bid_col": "V", "contract_col": "W"},
    ]
    col_map = {c: i for i, c in enumerate("ABCDEFGHIJKLMNOPQRSTUVWXYZ", 1)}

    current_month = None
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        cells = {}
        for c in row:
            try:
                cells[c.column_letter] = c.value
            except AttributeError:
                pass
        # Check for month headers
        for s in sections:
            val = cells.get(s["name_col"])
            if isinstance(val, str) and val.strip() in (
                "January", "February", "March", "April", "May", "June",
                "July", "August", "September", "October", "November", "December"
            ):
                current_month = val.strip()
                break

        for s in sections:
            name = cells.get(s["name_col"])
            if not name or not isinstance(name, str) or name.strip() in (
                "Name", "January", "February", "March", "April", "May", "June",
                "July", "August", "September", "October", "November", "December",
                "Total", "Totals"
            ):
                continue
            bid_amt = safe_float(cells.get(s["bid_col"]))
            contract_amt = safe_float(cells.get(s["contract_col"]))
            if bid_amt is None and contract_amt is None:
                continue
            job_num = cells.get("A") if s["name_col"] in ("B", "H") else cells.get("M") if s["name_col"] == "N" else None
            if isinstance(job_num, (int, float)):
                job_num = None
            bids.append({
                "year": 2026,
                "name": str(name).strip(),
                "job_number": str(job_num).strip() if job_num else None,
                "estimator": s["estimator"],
                "bid_type": s["type"],
                "bid_date": parse_date(cells.get(s["date_col"])),
                "bid_amount": bid_amt,
                "contract_amount": contract_amt,
                "won": contract_amt is not None and contract_amt > 0,
            })
    return bids


def parse_two_section(ws, year, left_label="Construction", right_label="Cabinets"):
    """2-section layout used in 2024, 2025, and similar years"""
    bids = []
    # Left section: B=Name, C=Date, D=Bid, E=Contract; Right: H=Name, I=Date, J=Bid, K=Contract
    sections = [
        {"label": left_label, "job_col": "A", "name_col": "B", "date_col": "C", "bid_col": "D" if year != 2023 else "E", "contract_col": "E" if year != 2023 else "F"},
        {"label": right_label, "job_col": "J" if year >= 2021 else "K", "name_col": "K" if year >= 2021 else "L", "date_col": "L" if year >= 2021 else "M", "bid_col": "M" if year >= 2021 else "N", "contract_col": "N" if year >= 2021 else None},
    ]
    # Adjust for 2023+ where cabinet section starts at col H or J
    if year >= 2024:
        sections[1] = {"label": right_label, "job_col": None, "name_col": "H", "date_col": "I", "bid_col": "J", "contract_col": "K"}

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        cells = {}
        for c in row:
            try:
                cells[c.column_letter] = c.value
            except AttributeError:
                pass
        for s in sections:
            name = cells.get(s["name_col"])
            if not name or not isinstance(name, str):
                continue
            skip = {"Name", "January", "February", "March", "April", "May", "June",
                    "July", "August", "September", "October", "November", "December",
                    "Total", "Totals", "MAC Construction Jobs", "MAC Cabinet Jobs", "MAC Jobs",
                    "Construction", "Cabinets"}
            if name.strip() in skip:
                continue
            bid_amt = safe_float(cells.get(s["bid_col"]))
            contract_amt = safe_float(cells.get(s["contract_col"])) if s["contract_col"] else None
            if bid_amt is None and contract_amt is None:
                continue
            job_num = cells.get(s["job_col"]) if s["job_col"] else None
            if isinstance(job_num, (int, float)):
                job_num = None
            bids.append({
                "year": year,
                "name": str(name).strip(),
                "job_number": str(job_num).strip() if job_num else None,
                "estimator": None,
                "bid_type": s["label"],
                "bid_date": parse_date(cells.get(s["date_col"])),
                "bid_amount": bid_amt,
                "contract_amount": contract_amt,
                "won": contract_amt is not None and contract_amt > 0,
            })
    return bids


# ── 2. Parse job cost sheets ──

def parse_job_cost_2024_2025(ws):
    """Job Cost 2024 & 2025: Estimator, Job #, Job Name, Final Amount, Materials, Labor, Total Costs, Gross Profit, Percent"""
    costs = []
    current_estimator = None
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        cells = {}
        for c in row:
            try:
                cells[c.column_letter] = c.value
            except AttributeError:
                pass
        # Check for estimator header rows
        est = cells.get("A")
        if isinstance(est, str) and est.strip() in ("Rusty", "Josh", "Daniel"):
            jnum = cells.get("B")
            if jnum and isinstance(jnum, str) and jnum.startswith("C"):
                current_estimator = est.strip()
            elif not cells.get("B"):
                current_estimator = est.strip()
                continue
        if isinstance(est, str) and "WIP" in est:
            continue

        job_num = cells.get("B")
        job_name = cells.get("C")
        if not job_num or not job_name:
            continue
        if isinstance(job_num, str) and not job_num.startswith("C"):
            if est and isinstance(est, str) and est.strip() in ("Rusty", "Josh", "Daniel"):
                current_estimator = est.strip()
            continue

        final_amt = safe_float(cells.get("D"))
        materials = safe_float(cells.get("E"))
        labor = safe_float(cells.get("F"))
        total_costs = safe_float(cells.get("G"))
        gross_profit = safe_float(cells.get("H"))
        profit_pct = safe_float(cells.get("I"))
        # Normalize: if stored as whole number (e.g. 25 for 25%), convert to fraction
        if profit_pct is not None and abs(profit_pct) > 1.0:
            profit_pct = profit_pct / 100.0

        if final_amt is None and materials is None:
            continue

        costs.append({
            "job_number": str(job_num).strip(),
            "job_name": str(job_name).strip(),
            "estimator": current_estimator,
            "final_amount": final_amt,
            "materials": materials,
            "labor": labor,
            "total_costs": total_costs,
            "gross_profit": gross_profit,
            "profit_percent": profit_pct,
            "source": "2024-2025",
        })
    return costs


def parse_job_cost_legacy(ws):
    """Job Cost (2018-2023): Job #, Job Name, Labor, Burden 30%, material, Cost, Contract amount, PO, %, Estimator"""
    costs = []
    current_year = None
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        cells = {}
        for c in row:
            try:
                cells[c.column_letter] = c.value
            except AttributeError:
                pass
        a_val = cells.get("A")

        # Year markers
        if isinstance(a_val, (int, float)) and 2018 <= a_val <= 2025:
            current_year = int(a_val)
            continue

        if not a_val or not isinstance(a_val, str) or not a_val.startswith("C"):
            continue

        job_name = cells.get("B")
        if not job_name:
            continue

        labor = safe_float(cells.get("C"))
        burden = safe_float(cells.get("D"))
        material = safe_float(cells.get("E"))
        cost = safe_float(cells.get("F"))
        contract_amt = safe_float(cells.get("G"))
        profit_pct = safe_float(cells.get("I"))
        # Normalize: if stored as whole number (e.g. 25 for 25%), convert to fraction
        if profit_pct is not None and abs(profit_pct) > 1.0:
            profit_pct = profit_pct / 100.0
        estimator = cells.get("J")

        if cost is None and contract_amt is None:
            continue

        gross_profit = (contract_amt - cost) if (contract_amt and cost) else None

        costs.append({
            "job_number": str(a_val).strip(),
            "job_name": str(job_name).strip(),
            "estimator": str(estimator).strip() if estimator else None,
            "final_amount": contract_amt,
            "materials": material,
            "labor": (labor or 0) + (burden or 0) if labor else labor,
            "total_costs": cost,
            "gross_profit": gross_profit,
            "profit_percent": profit_pct,
            "source": f"legacy-{current_year}" if current_year else "legacy",
        })
    return costs


# ── 3. Parse individual bid xlsx files from P:/Mac2026/Cabinets/Bids ──

def parse_bid_xlsx(filepath):
    """Parse a bid template xlsx to extract line items, contact, and total.
    Template layout is consistent: rows 3-5 MAC header, row 10 date, row 11 contact,
    row 12 contractor, row 13 email, row 14 phone, row 16 RE:, then rooms+items."""
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb[wb.sheetnames[0]]
    except Exception:
        return None

    def get_cells(row_data):
        cells = {}
        for c in row_data:
            try:
                if c.value is not None:
                    cells[c.column_letter] = c.value
            except AttributeError:
                pass
        return cells

    # Read all rows up front
    all_rows = list(ws.iter_rows(min_row=1, max_row=min(300, ws.max_row), values_only=False))

    result = {
        "file": os.path.basename(filepath),
        "contact_name": None,
        "contractor": None,
        "email": None,
        "phone": None,
        "job_name": None,
        "date": None,
        "rooms": [],
        "total": None,
    }

    # Phase 1: Parse header (first 20 rows) positionally
    header_done = False
    data_start = 0
    for i, row in enumerate(all_rows[:25]):
        cells = get_cells(row)
        a = cells.get("A")
        d = cells.get("D")
        a_str = str(a).strip() if a else ""

        # Date in column D
        if isinstance(d, (datetime, date)) and not result["date"]:
            result["date"] = parse_date(d)
            # Next rows should be: contact, contractor, email, phone, blank, RE:
            for j in range(1, 8):
                if i + j >= len(all_rows):
                    break
                nc = get_cells(all_rows[i + j])
                na = nc.get("A")
                na_str = str(na).strip() if na else ""
                if not na:
                    continue
                if not result["contact_name"] and na and "@" not in na_str and not na_str.startswith("RE:") and not na_str.startswith("816") and not na_str.startswith("913"):
                    result["contact_name"] = na_str
                elif not result["contractor"] and na and "@" not in na_str and not na_str.startswith("RE:") and not na_str.startswith("816") and not na_str.startswith("913"):
                    result["contractor"] = na_str
                elif "@" in na_str and not result["email"]:
                    result["email"] = na_str
                elif re.match(r"^\d{3}[\-\.]", na_str) and not result["phone"]:
                    result["phone"] = na_str
                elif na_str.startswith("RE:"):
                    result["job_name"] = na_str[3:].strip()
                    data_start = i + j + 1
                    header_done = True
                    break
            break

    if not header_done:
        # Try to find RE: line as fallback
        for i, row in enumerate(all_rows[:25]):
            cells = get_cells(row)
            a = cells.get("A")
            if a and str(a).strip().startswith("RE:"):
                result["job_name"] = str(a).strip()[3:].strip()
                data_start = i + 1
                break

    # Phase 2: Parse rooms and line items
    current_room = None
    line_items = []

    for row in all_rows[data_start:]:
        cells = get_cells(row)
        a = cells.get("A")
        b = cells.get("B")
        c_val = cells.get("C")
        d = cells.get("D")

        if not a and not b and not c_val and not d:
            continue

        a_str = str(a).strip() if a else ""

        # Total line
        if a_str.lower().startswith("total casework"):
            result["total"] = safe_float(d)
            if current_room and line_items:
                result["rooms"].append({"name": current_room, "items": line_items})
            current_room = None
            line_items = []
            continue

        # Boilerplate - stop parsing
        if a_str.startswith("Standard MAC") or a_str.startswith("Submitted By"):
            break

        # Line item (has name in A, quantity in B, price in C)
        if a and b is not None and c_val is not None:
            qty = safe_float(b)
            price = safe_float(c_val)
            if qty is not None and price is not None and price > 0:
                line_items.append({
                    "name": a_str,
                    "quantity": qty,
                    "price": price,
                    "unit_price": round(price / qty, 2) if qty > 0 else 0,
                })
                continue

        # Room/area header (text in A only, no numeric data)
        if a and b is None and c_val is None and (d is None or d == 0):
            if not a_str.startswith("**") and not a_str.startswith("Excludes") and len(a_str) < 80:
                if current_room and line_items:
                    result["rooms"].append({"name": current_room, "items": line_items})
                    line_items = []
                current_room = a_str

    # Flush remaining
    if current_room and line_items:
        result["rooms"].append({"name": current_room, "items": line_items})

    if not result["rooms"] and not result["total"] and not result["contractor"]:
        return None
    return result


# ── 4. Build room templates from detailed bids ──

def normalize_room_name(name):
    """Normalize a room name by stripping room numbers, multiplier text, and standardizing common names."""
    n = name.strip()
    # Strip leading room/suite numbers like "112 ", "A-101 ", "Suite 200 "
    n = re.sub(r'^(?:Suite\s+|Room\s+|Rm\s+)?[A-Z]?-?\d{1,4}[A-Z]?\s+', '', n)
    # More targeted: strip leading digits/letters that look like room numbers
    n = re.sub(r'^[A-Z]?\d{1,4}[A-Z]?\s+[-–]\s*', '', n)
    n = re.sub(r'^[A-Z]?\d{1,4}[A-Z]?\s+', '', n)
    # Strip multiplier text like "(QTY: 6)" or "(x4)"
    n = re.sub(r'\s*\(QTY:\s*\d+\)', '', n, flags=re.IGNORECASE)
    n = re.sub(r'\s*\(x\d+\)', '', n, flags=re.IGNORECASE)
    # Strip trailing numbers/letters in parens like "(A)" or "(1)"
    n = re.sub(r'\s*\([A-Z0-9]\)$', '', n)
    # Normalize common variations
    n = n.strip()
    replacements = {
        "Breakroom": "Break Room",
        "Break room": "Break Room",
        "breakroom": "Break Room",
        "break room": "Break Room",
        "Restrooms": "Restroom",
        "Rest Room": "Restroom",
        "Mens Restroom": "Restroom",
        "Womens Restroom": "Restroom",
        "Men's Restroom": "Restroom",
        "Women's Restroom": "Restroom",
        "Men's Room": "Restroom",
        "Women's Room": "Restroom",
        "Kitchenette": "Kitchen",
    }
    for k, v in replacements.items():
        if n.lower() == k.lower():
            return v
    # Title-case normalize
    if n:
        n = n[0].upper() + n[1:]
    return n


def build_room_templates(detailed_bids):
    """Build room template data from all detailed bids.
    Returns dict of normalized room name -> {count, items: {item_name: {avg_qty, avg_price, count}}}
    Only includes rooms with 3+ occurrences.
    """
    room_data = defaultdict(lambda: {"occurrences": [], "items": defaultdict(list)})

    for bid in detailed_bids:
        for room in bid["rooms"]:
            norm = normalize_room_name(room["name"])
            if not norm or len(norm) < 2:
                continue
            room_data[norm]["occurrences"].append(bid["file"])
            for item in room["items"]:
                room_data[norm]["items"][item["name"]].append({
                    "quantity": item["quantity"],
                    "unit_price": item["unit_price"],
                })

    # Filter to rooms with 3+ occurrences and build stats
    templates = {}
    for name, data in room_data.items():
        count = len(set(data["occurrences"]))  # unique bids
        if count < 3:
            continue
        items = {}
        for item_name, entries in data["items"].items():
            if len(entries) < 2:
                continue
            qtys = [e["quantity"] for e in entries if e["quantity"] and e["quantity"] > 0]
            prices = [e["unit_price"] for e in entries if e["unit_price"] and e["unit_price"] > 0]
            if not qtys:
                continue
            items[item_name] = {
                "avg_qty": round(sum(qtys) / len(qtys), 1),
                "avg_price": round(sum(prices) / len(prices), 2) if prices else 0,
                "count": len(entries),
            }
        if items:
            templates[name] = {
                "count": count,
                "items": items,
            }

    return templates


# ── Main ──

def main():
    print("Loading bid sheet...")
    wb = openpyxl.load_workbook(BID_SHEET, data_only=True)

    all_bids = []
    all_costs = []
    all_detailed_bids = []

    # Parse yearly bid tracking
    print("Parsing 2026 bids...")
    all_bids.extend(parse_2026(wb["2026"]))

    for year in [2025, 2024]:
        print(f"Parsing {year} bids...")
        all_bids.extend(parse_two_section(wb[str(year)], year))

    for year in [2023, 2022, 2021, 2020, 2019]:
        if str(year) in wb.sheetnames:
            print(f"Parsing {year} bids...")
            all_bids.extend(parse_two_section(wb[str(year)], year))

    # Parse job costs
    print("Parsing job costs 2024-2025...")
    all_costs.extend(parse_job_cost_2024_2025(wb["Job Cost 2024 & 2025"]))

    print("Parsing legacy job costs...")
    all_costs.extend(parse_job_cost_legacy(wb["Job Cost"]))

    # Parse detailed bids from P:/
    print(f"Parsing detailed bids from {BIDS_DIR}...")
    if os.path.exists(BIDS_DIR):
        for folder in os.listdir(BIDS_DIR):
            fpath = os.path.join(BIDS_DIR, folder)
            if not os.path.isdir(fpath):
                continue
            for f in os.listdir(fpath):
                if f.endswith(".xlsx") and not f.startswith("~"):
                    result = parse_bid_xlsx(os.path.join(fpath, f))
                    if result:
                        all_detailed_bids.append(result)

    # ── Build room templates ──
    print("\nBuilding room templates...")
    room_templates = build_room_templates(all_detailed_bids)
    print(f"Found {len(room_templates)} room template types")
    for name, data in sorted(room_templates.items(), key=lambda x: -x[1]["count"]):
        print(f"  {name}: {data['count']} occurrences, {len(data['items'])} item types")

    # ── Build analytics ──
    print("\nBuilding analytics...")

    # Pricing stats by line item type
    item_prices = defaultdict(list)
    for bid in all_detailed_bids:
        for room in bid["rooms"]:
            for item in room["items"]:
                # Normalize item name
                name = item["name"]
                for prefix in ["PLAM ", "Paint Grade ", "Stain Grade "]:
                    if name.startswith(prefix):
                        name = name[len(prefix):]
                        break
                item_prices[name].append({
                    "unit_price": item["unit_price"],
                    "quantity": item["quantity"],
                    "total": item["price"],
                    "job": bid["job_name"],
                    "date": bid["date"],
                })

    pricing_stats = {}
    for name, prices in sorted(item_prices.items()):
        if len(prices) < 2:
            continue
        unit_prices = [p["unit_price"] for p in prices if p["unit_price"] > 0]
        if not unit_prices:
            continue
        pricing_stats[name] = {
            "count": len(prices),
            "avg_unit_price": round(sum(unit_prices) / len(unit_prices), 2),
            "min_unit_price": round(min(unit_prices), 2),
            "max_unit_price": round(max(unit_prices), 2),
            "median_unit_price": round(sorted(unit_prices)[len(unit_prices) // 2], 2),
            "total_revenue": round(sum(p["total"] for p in prices), 2),
        }

    # Win rate analysis
    # Filter to cabinet bids only (exclude Construction)
    cab_bids = [b for b in all_bids if b.get("bid_type") != "Construction"]

    won = [b for b in cab_bids if b["won"]]
    lost = [b for b in cab_bids if not b["won"] and b["bid_amount"]]
    win_rate = len(won) / max(len(won) + len(lost), 1)

    # Win rate by year
    win_by_year = defaultdict(lambda: {"won": 0, "lost": 0, "revenue": 0})
    for b in cab_bids:
        y = b["year"]
        if b["won"]:
            win_by_year[y]["won"] += 1
            win_by_year[y]["revenue"] += b["contract_amount"] or 0
        elif b["bid_amount"]:
            win_by_year[y]["lost"] += 1

    # Contractor rankings
    contractor_stats = defaultdict(lambda: {"jobs": 0, "revenue": 0, "bids": []})
    for bid in all_detailed_bids:
        if bid["contractor"]:
            c = bid["contractor"]
            contractor_stats[c]["jobs"] += 1
            contractor_stats[c]["revenue"] += bid["total"] or 0
            contractor_stats[c]["bids"].append(bid["job_name"])

    # Also from bid tracking
    for b in all_bids:
        if b["won"] and b["contract_amount"]:
            # Use name as proxy since we don't have contractor in bid sheet
            pass

    contractor_leaderboard = sorted(
        [{"name": k, **{kk: vv for kk, vv in v.items() if kk != "bids"}, "recent_jobs": v["bids"][-5:]}
         for k, v in contractor_stats.items() if v["jobs"] >= 1],
        key=lambda x: x["revenue"], reverse=True
    )

    # Job cost profitability — cabinet shop only (2024-2025 source = Rusty/cabinets)
    cabinet_costs = [c for c in all_costs if c["source"] == "2024-2025"]
    profitable = [c for c in cabinet_costs if c["profit_percent"] is not None]
    avg_margin = sum(c["profit_percent"] for c in profitable) / max(len(profitable), 1)

    # Normalize estimator names (legacy data has job numbers mixed in)
    estimator_map = {}
    for c in all_costs:
        if c["estimator"]:
            raw = c["estimator"]
            # Extract just the name part
            name = raw.split("/")[0].split(" - ")[0].strip()
            # Remove trailing job numbers
            name = re.sub(r"\s+\d{4,}.*$", "", name)
            name = re.sub(r"\s+[CR]\d+.*$", "", name)
            if name and not name.replace(".", "").isdigit():
                estimator_map[raw] = name
                c["estimator"] = name
            else:
                c["estimator"] = None

    margin_by_estimator = defaultdict(list)
    for c in profitable:
        if c["estimator"]:
            margin_by_estimator[c["estimator"]].append(c["profit_percent"])

    # Seasonal analysis (cabinets only)
    monthly_bids = defaultdict(lambda: {"count": 0, "total_value": 0})
    for b in cab_bids:
        if b["bid_date"]:
            try:
                month = int(b["bid_date"].split("-")[1])
                monthly_bids[month]["count"] += 1
                monthly_bids[month]["total_value"] += b["bid_amount"] or 0
            except (IndexError, ValueError):
                pass

    # Client/contact lookup for auto-fill — from all detailed bid files (incl. those with only header data)
    contacts = {}
    for bid in all_detailed_bids:
        if bid["contractor"] and bid["contractor"] not in contacts:
            contacts[bid["contractor"]] = {
                "name": bid["contact_name"],
                "email": bid["email"],
                "phone": bid.get("phone"),
            }
        elif bid["contractor"] and bid["contractor"] in contacts:
            # Fill in missing fields from other bids
            existing = contacts[bid["contractor"]]
            if not existing.get("name") and bid["contact_name"]:
                existing["name"] = bid["contact_name"]
            if not existing.get("email") and bid["email"]:
                existing["email"] = bid["email"]
            if not existing.get("phone") and bid.get("phone"):
                existing["phone"] = bid.get("phone")

    # Estimator stats from bid data (covers Josh + anyone without cost data)
    # Pre-2025 bids with no estimator are attributed to Rusty
    def get_bid_estimator(b):
        if b.get("estimator"):
            return b["estimator"]
        if b.get("year", 0) < 2025:
            return "Rusty"
        return None

    bid_estimator_stats = defaultdict(lambda: {"bids": 0, "won": 0, "revenue_won": 0})
    for b in cab_bids:
        est = get_bid_estimator(b)
        if est:
            bid_estimator_stats[est]["bids"] += 1
            if b["won"]:
                bid_estimator_stats[est]["won"] += 1
                bid_estimator_stats[est]["revenue_won"] += b.get("contract_amount") or b["bid_amount"] or 0

    # Win probability model — multi-factor: amount + year + estimator + bid type
    # Excludes 2020 (bad data — no contract amounts)
    valid_bids = [b for b in cab_bids if b.get("year", 0) != 2020 and b.get("bid_amount")]
    valid_won = sum(1 for b in valid_bids if b["won"])
    base_rate = valid_won / max(len(valid_bids), 1)

    # Amount buckets
    amount_breaks = [(0, 5000), (5000, 15000), (15000, 30000), (30000, 50000), (50000, 100000), (100000, 999999)]
    win_prob_table = []
    for lo, hi in amount_breaks:
        w = sum(1 for b in valid_bids if b["won"] and lo <= b["bid_amount"] < hi)
        t = sum(1 for b in valid_bids if lo <= b["bid_amount"] < hi)
        win_prob_table.append({
            "min_amount": lo, "max_amount": hi,
            "win_rate": round(w / max(t, 1), 3), "sample_size": t
        })

    # Year adjustments
    year_adj = {}
    for y in range(2021, 2027):
        w = sum(1 for b in valid_bids if b["won"] and b.get("year") == y)
        t = sum(1 for b in valid_bids if b.get("year") == y)
        if t >= 10:
            year_adj[str(y)] = round((w / t) / max(base_rate, 0.01), 3)

    # Estimator adjustments
    # Pre-2025 bids have no estimator tag but are ALL Rusty's
    def get_estimator(b):
        if b.get("estimator"):
            return b["estimator"]
        if b.get("year", 0) < 2025:
            return "Rusty"
        return None

    est_adj = {}
    for est in ["Rusty", "Josh"]:
        w = sum(1 for b in valid_bids if b["won"] and get_estimator(b) == est)
        t = sum(1 for b in valid_bids if get_estimator(b) == est)
        if t >= 5:
            est_adj[est] = round((w / t) / max(base_rate, 0.01), 3)

    # Bid type adjustments
    type_adj = {}
    for bt in ["Third Party", "Internal"]:
        w = sum(1 for b in valid_bids if b["won"] and b.get("bid_type") == bt)
        t = sum(1 for b in valid_bids if b.get("bid_type") == bt)
        if t >= 5:
            type_adj[bt] = round((w / t) / max(base_rate, 0.01), 3)

    # ── Assemble output ──
    output = {
        "generated": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "summary": {
            "total_bids_tracked": len(cab_bids),
            "total_jobs_costed": len(cabinet_costs),
            "total_detailed_bids": len(all_detailed_bids),
            "overall_win_rate": round(win_rate, 3),
            "avg_profit_margin": round(avg_margin, 3) if profitable else None,
            "years_covered": sorted(set(b["year"] for b in cab_bids)),
        },
        "pricing": pricing_stats,
        "win_rates": {
            "overall": round(win_rate, 3),
            "by_year": {str(k): {"won": v["won"], "lost": v["lost"],
                                  "rate": round(v["won"] / max(v["won"] + v["lost"], 1), 3),
                                  "revenue": round(v["revenue"], 2)}
                        for k, v in sorted(win_by_year.items())},
        },
        "win_probability": {
            "base_rate": round(base_rate, 3),
            "by_amount": win_prob_table,
            "year_adjustments": year_adj,
            "estimator_adjustments": est_adj,
            "type_adjustments": type_adj,
        },
        "contractors": contractor_leaderboard[:30],
        "profitability": {
            "avg_margin": round(avg_margin, 3) if profitable else None,
            "by_estimator": {k: {"avg": round(sum(v) / len(v), 3), "count": len(v)}
                            for k, v in margin_by_estimator.items()},
            "bid_estimators": {k: {"bids": v["bids"], "won": v["won"], "revenue_won": round(v["revenue_won"], 2),
                                    "win_rate": round(v["won"] / max(v["bids"], 1), 3)}
                               for k, v in bid_estimator_stats.items()},
            "top_jobs": sorted(
                [{"job": c["job_number"], "name": c["job_name"], "margin": c["profit_percent"],
                  "revenue": c["final_amount"],
                  "profit": round((c["final_amount"] or 0) * (c["profit_percent"] or 0), 2)}
                 for c in profitable if c["final_amount"] and c["final_amount"] > 1000
                 and c["job_number"].startswith("C2")
                 and int(re.sub(r'\D', '', c["job_number"][1:]) or "0") >= 2100
                 and (c["profit_percent"] or 0) > 0],
                key=lambda x: x["profit"], reverse=True
            )[:20],
            "worst_jobs": sorted(
                [{"job": c["job_number"], "name": c["job_name"], "margin": c["profit_percent"],
                  "revenue": c["final_amount"],
                  "profit": round((c["final_amount"] or 0) * (c["profit_percent"] or 0), 2)}
                 for c in profitable if c["final_amount"] and c["final_amount"] > 1000
                 and c["job_number"].startswith("C2")
                 and int(re.sub(r'\D', '', c["job_number"][1:]) or "0") >= 2100],
                key=lambda x: x["profit"]
            )[:20],
        },
        "seasonal": {str(k): v for k, v in sorted(monthly_bids.items())},
        "contacts": contacts,
        "room_templates": room_templates,
        "job_costs": cabinet_costs,
        "bids": cab_bids,
    }

    # Write output
    os.makedirs(os.path.dirname(OUTPUT), exist_ok=True)
    with open(OUTPUT, "w") as f:
        json.dump(output, f, indent=2, default=str)

    # Print summary
    print(f"\n{'='*60}")
    print(f"Extracted {len(cab_bids)} cabinet bids, {len(cabinet_costs)} cabinet job costs, {len(all_detailed_bids)} detailed bids")
    print(f"Win rate: {win_rate:.1%}")
    if profitable:
        print(f"Avg profit margin: {avg_margin:.1%}")
    print(f"\nPricing stats for {len(pricing_stats)} line item types:")
    for name, stats in list(pricing_stats.items())[:10]:
        print(f"  {name}: ${stats['avg_unit_price']}/unit (${stats['min_unit_price']}-${stats['max_unit_price']}, n={stats['count']})")
    print(f"\nTop contractors:")
    for c in contractor_leaderboard[:10]:
        print(f"  {c['name']}: {c['jobs']} jobs, ${c['revenue']:,.0f}")
    print(f"\nSeasonal (bids by month):")
    for m, v in sorted(monthly_bids.items()):
        bar = "#" * (v["count"] // 5)
        print(f"  {m:2d}: {v['count']:3d} bids ${v['total_value']:>12,.0f} {bar}")
    print(f"\nOutput: {OUTPUT}")


if __name__ == "__main__":
    main()
